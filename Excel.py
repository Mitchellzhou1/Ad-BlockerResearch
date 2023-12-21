## need to convert to CSV later on

import openpyxl
import csv

wb = openpyxl.Workbook()
ws = wb.active

row = 2
intercept_row = 2
timeout_row = 2
other_row = 2
noscan_row = 2
notInteractable_row = 2
(intercept_lst, timeout_lst, notInteractable_lst,
 staleElems_lst, noScan_lst, other_lst) = [[] for i in range(6)]
HTML_obj = "None"

def initialize():
    ws['A1'] = "Dropdown Opened?"
    ws['B1'] = "Outer HTML Change"
    ws['C1'] = "DOM structure Change"
    ws['D1'] = "Initial Outer HTML "
    ws['E1'] = "After Click Outer HTML"
    ws['F1'] = "Initial DOM Structure"
    ws['G1'] = "After Click DOM Structure"
    ws['H1'] = "Initial Link"
    ws['I1'] = "After Click Link"
    ws['J1'] = "Tries"

    ws['L1'] = "Links with Intercept Errors:"
    ws['M1'] = "Links with Timeout Errors:"
    ws['N1'] = "Links with Not Interactable Exception:"
    ws['O1'] = "Links could not be scanned:"
    ws['P1'] = "Links with Stale Element:"
    ws['Q1'] = "Links with unknown Errors:"
    wb.save("TESTING.xlsx")


def write_intercepts(site):
    global intercept_row
    if site not in intercept_lst:
        intercept_lst.append(site)
        ws[f'G{intercept_row}'] = site
        intercept_row += 1
        wb.save("TESTING.xlsx")


def write_timeout_row(site):
    global timeout_row
    if site not in timeout_lst:
        timeout_lst.append(site)
        ws[f'H{timeout_row}'] = site
        timeout_row += 1
        wb.save("TESTING.xlsx")


def write_notInteractable_row(site):
    global notInteractable_row
    if site not in notInteractable_lst:
        notInteractable_lst.append(site)
        ws[f'I{noscan_row}'] = site
        notInteractable_row += 1
        wb.save("TESTING.xlsx")


def write_noscan_row(site):
    global noscan_row
    if site not in noScan_lst:
        noScan_lst.append(site)
        ws[f'J{noscan_row}'] = site
        noscan_row += 1
        wb.save("TESTING.xlsx")


def write_other_row(site):
    global other_row
    if site not in other_lst:
        other_lst.append(site)
        ws[f'K{other_row}'] = site
        other_row += 1
        wb.save("TESTING.xlsx")


def write_test_size(tested, total):
    global row
    row -= 1
    ws[f'B{row}'] = f"{tested} / {total}"
    row += 1


def write_results(data):
    global row

    # Format = ['True redirect', HTML / link Before, HTML / link After, tries, DOM]
    if type(data) is str:  # Website
        row += 1
        ws[f'A{row}'] = data  # website

    elif len(data) == 5:  # Errors
        ws[f'A{row}'] = data[0]  # Error Message
        ws[f'B{row}'] = data[1]  # OuterHTML Change
        ws[f'C{row}'] = data[2]  # DOM Change
        ws[f'D{row}'] = data[3]  # Button HTML
        ws[f'J{row}'] = f'Tries: {data[4]}'  # number of Tries

    elif len(data) == 10:  # Clicked button results
        ws[f'A{row}'] = data[0]  # opened?
        ws[f'B{row}'] = data[1]  # OuterHTML Change
        ws[f'C{row}'] = data[2]  # DOM Change
        ws[f'D{row}'] = data[3]  # InitialOuterHTML
        ws[f'E{row}'] = data[4]  # AfterClickOuterHTML
        ws[f'F{row}'] = data[5]  # DOM Initial
        ws[f'G{row}'] = data[6]  # DOM After Click
        ws[f'H{row}'] = data[7]  # URL Initial
        ws[f'I{row}'] = data[8]  # URL AFter
        ws[f'J{row}'] = f'Tries: {data[9]}'  # number of Tries

    row += 1
    wb.save("TESTING.xlsx")


def initialize_csv_file(HTML_obj):
    filename = f"{HTML_obj}.csv"
    fieldnames = ["Website Name", "HTML Content"]

    with open(filename, mode='w', newline='', encoding='utf-8') as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()


def add_to_csv(website_name, html_content, HTML_obj):
    filename = f"{HTML_obj}.csv"
    object = f"{HTML_obj} HTML Content"
    fieldnames = ["Website Name", object]

    data = {"Website Name": website_name, object: html_content}

    with open(filename, mode='a', newline='', encoding='utf-8') as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)

        # Write data rows
        writer.writerow(data)

    print("Added Elem")


