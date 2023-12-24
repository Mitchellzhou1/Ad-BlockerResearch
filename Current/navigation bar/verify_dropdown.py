from selenium import webdriver
from selenium.common import ElementClickInterceptedException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException
from time import *
from bs4 import BeautifulSoup
import requests

import openpyxl
import pyautogui
import signal
from tranco import Tranco

# Prepare Chrome
options = Options()
# options.headless = False
# options.add_argument("--headless=new")
options.add_argument("--no-sandbox")
options.add_argument("--disable-animations")
options.add_argument("--disable-web-animations")
# options.add_argument("--incognito")
# options.add_argument("--single-process")
options.add_argument("--disable-gpu")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--disable-web-security")
options.add_argument("--disable-features=IsolateOrigins,site-per-process")
options.add_argument("--disable-features=AudioServiceOutOfProcess")
# options.add_argument("auto-open-devtools-for-tabs")
options.add_argument(
    "user-agent=Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36")
# options.add_extension("/home/seluser/measure/harexporttrigger-0.6.3.crx")

adBlockerIDs = {"adblockPlus": 'cfhdojbkjhnklbpkdaibdccddilifddb'}

sites = [
    'https://en.wikipedia.org/wiki/Main_Page',
    'https://www.amazon.com/',
    'https://www.microsoft.com/en-us/',
    'https://www.office.com/',
    'https://weather.com/',
    'https://openai.com/',
    'https://www.bing.com/',
    'https://duckduckgo.com/',
    'https://cnn.com',
    'https://www.nytimes.com/',
    'https://www.twitch.tv/',
    'https://www.imdb.com/',
    'https://mail.ru/',
    'https://naver.com',
    'https://zoom.us/',
    'https://www.globo.com/',
    'https://www.ebay.com/',
    'https://www.foxnews.com/',
    'https://www.instructure.com/',
    'https://www.walmart.com/',
    'https://www.indeed.com/',
    'https://www.paypal.com/us/home',
    'https://www.accuweather.com/',
    'https://www.pinterest.com/',
    'https://www.bbc.com/',
    'https://www.homedepot.com/',
    'https://www.breitbart.com/',
    'https://github.com/'
]

tag = ['button',
       'div',
       'input',
       'svg',
       'a'
       ]

attributes = [
    'false',
    'true',
    'main menu',
    'open menu',
    'all microsoft menu',
    'menu',
    'navigation',
    'primary navigation',
    'hamburger',
    'settings and quick links',
    'dropdown',
    'dialog',
    'js-menu-toggle',
    'searchDropdownDescription',
    'ctabutton',
    'legacy-homepage_legacyButton__oUMB9 legacy-homepage_hamburgerButton__VsG7q',
    'Toggle language selector',
    'Open Navigation Drawer',
    'guide',
    'Expand Your Library',
    'Collapse Your Library'
]

xpaths = [
    '@aria-expanded',
    '@aria-label',
    '@class',
    '@aria-haspopup',
    '@aria-describedby',
    '@data-testid',
]

driver = None
icon = 0

class TimeoutError(Exception):
    pass


def signal_handler(signum, frame):
    raise TimeoutError("Function execution time exceeded the limit")


def initialize(adblocker, seconds=14):
    """
    This function will start a Chrome instance with the option of installing an ad blocker.
    Adjust the seconds parameter so that it will wait for the ad blocker to finish downloading.
    """
    chrome_options = webdriver.ChromeOptions()

    if adblocker:
        chrome_options.add_extension('adBlockerPlus.crx')

    chrome_options.add_argument('--disable-dev-shm-usage')
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

    # give it time to install
    if adblocker:
        sleep(seconds)
        pyautogui.hotkey('ctrl', 'w')

    return driver



def load_site(url, skipped=[]):
    # url = f'https://{url}'
    try:
        response = requests.get(url)
        if response.status_code == 200:
            driver.get(url)
            wait = WebDriverWait(driver, 15)  # Changed timeout to 15 seconds
            try:
                wait.until(EC.presence_of_element_located((By.XPATH, "//*")))
                sleep(2)
                return True
            except TimeoutException:
                raise TimeoutError("Took too long to load...")
    except Exception:
        skipped.append(url)
        return False


def count_tags():
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    tags = soup.find_all()
    return len(tags)


def find_dropdown():
    def collect():
        found_elements = []
        for attribute in attributes:
            for path in xpaths:
                xpath = f'//*[translate({path}, "ABCDEFGHIJKLMNOPQRSTUVWXYZ", "abcdefghijklmnopqrstuvwxyz")="{attribute.lower()}"]'
                try:
                    elements = driver.find_elements(By.XPATH, xpath)
                    for element in elements:
                        if element not in found_elements:
                            found_elements.append(element)
                except Exception as e:
                    print(e)

        if len(found_elements) > 1:
            found_elements.sort(key=lambda e: driver.execute_script(
                "var elem = arguments[0], parents = 0; while (elem && elem.parentElement) { elem = elem.parentElement; parents++; } return parents;",
                e
            ))

        return found_elements

    try:
        ret = collect()
        if ret:
            return ret
    except Exception:
        pass

    sleep(5)
    return collect()


def cursor_change(element):
    actions = ActionChains(driver)
    try:
        actions.move_to_element(element).perform()
        cursor_property = element.value_of_css_property('cursor')
        if cursor_property == 'pointer':
            return True
        else:
            return False
    except Exception:
        if element.is_displayed():
            return True
        else:
            return False


def print_found_elems(lst):
    for i in lst:
        outer_html = i.get_attribute('outerHTML')
        first_line = outer_html.splitlines()[0]
        print(first_line)
        # print(i)


def check_redirect(url):
    def are_urls_equal(url1, url2):
        path1 = url1.rstrip('/').strip('https://').strip('www.')
        path2 = url2.rstrip('/').strip('https://').strip('www.')
        return path1 == path2

    sleep(3)
    if not are_urls_equal(driver.current_url, url):
        return True, driver.current_url

    all_windows = driver.window_handles
    if len(all_windows) > 1:
        for window in all_windows[1:]:
            driver.switch_to.window(window)
            driver.close()
        driver.switch_to.window(all_windows[0])
        return True, driver.current_url
    return False, driver.current_url


def write_data_to_file(data, intercept, timeout, other):
    def write_rows(lst, column, row):
        for i in range(len(lst)):
            ws[f'{column}{i+row}'] = lst[i]

    def set_up_errors(intercept, timeout, other):
        ws['A1'] = "Intercept Error"
        ws['B1'] = "Timeout Error"
        ws['c1'] = "Some Other Error"
        write_rows(intercept, 'A', 2)
        write_rows(timeout, 'B',2)
        write_rows(other, 'C', 2)
        return max(len(intercept), len(timeout), len(other))

    wb = openpyxl.Workbook()
    ws = wb.active
    row = set_up_errors(intercept, timeout, other) + 4
    ws[f'A{row}'] = "Results"
    ws[f'B{row}'] = "HTML / Link Before"
    ws[f'C{row}'] = "HTML / Link After"

    offset = row + 1
    # ['True redirect', HTML / link Before, HTML / link After, tries]

    for i in range(len(data)):
        if len(data[i]) == 1:
            ws[f'A{offset}'] = data[i][0]                       #website
            offset += 1
        else:
            ws[f'A{offset}'] = data[i][0]                       #result
            ws[f'B{offset}'] = data[i][1]                       #before
            ws[f'C{offset}'] = data[i][2]                       #after
            ws[f'D{offset}'] = f'Tries: {data[i][3]}'           #number of Tries
            offset += 1


    wb.save("buttons.xlsx")



########################################################################################################################
def intercept_handler(curr, icon, url):
    # def click_corners():
    #     screen_width = driver.execute_script("return window.innerWidth;")
    #     screen_height = driver.execute_script("return window.innerHeight;")
    #
    #     actions = ActionChains(driver)
    #
    #     # print("Click on the bottom right corner")
    #     actions.move_by_offset(screen_width - 1, screen_height - 1)
    #     actions.click()
    #     actions.perform()
    #
    #     # print("Click on the bottom left corner")
    #     actions.move_by_offset(1 - screen_width, screen_height - 1)
    #     actions.click()
    #     actions.perform()

    # try:
    #     click_corners()
    #     curr[icon].click()
    #     check_redirect(url)
    # except ElementClickInterceptedException:
    try:
        curr[icon - 1].click()
    except Exception:
        load_site(url)
        print("couldn't resolve intercept error")
        return find_dropdown(), icon

    return curr, icon - 1


def check_opened(url, button, initial_html, initial_tag):
    def check_HTML(initial, after):
        if initial != after:
            return True
        return False

    redirect, new = check_redirect(url)
    if redirect:
        return "True - redirect", new

    after_outer_html = button.get_attribute('outerHTML')
    clicked = after_outer_html.splitlines()[0]

    if check_HTML(initial_html, clicked):
        return "True - OuterHTML change", after_outer_html

    if count_tags() > initial_tag:
        return "True - More tags were generated", after_outer_html
    return "False", after_outer_html


def test_drop_down(curr, url, tries=1):
    # attempts to click the button and refreshes afterward
    global driver, icon
    signal.signal(signal.SIGALRM, signal_handler)
    signal.alarm(100)

    ret = []
    while icon < len(curr):
        try:
            outer_html = curr[icon].get_attribute('outerHTML')
            initial_html = outer_html.splitlines()[0]

        except Exception as e:
            icon += 1
            continue

        if not cursor_change(curr[icon]):
            icon += 1
            continue

        initial_tag = count_tags()
        curr[icon].click()

        check, after = check_opened(url, curr[icon], initial_html, initial_tag)
        # if check == "False":
        #     raise InterruptedError

        load_site(url)
        curr = find_dropdown()

        if check == "True - redirect":
            ret.append([check, url, after, tries])
        else:
            ret.append([check, outer_html, after, tries])

        icon += 1

    return ret


def main():
    t = Tranco(cache=True, cache_dir='.tranco')
    latest_list = t.list()
    sites = latest_list.top(10000)
    for i in sites:
        print(i)

main()

while 1:
    1



# fastly.net: this website is broken
# yahoo.com: going on it on crawler is different then on browser
#  bit.ly   Message: no such execution context
# msn.com
# yandex.net --robot detection

#naver.com: says its redirect
#cnn.com
#golbo.com
#softonic.com
