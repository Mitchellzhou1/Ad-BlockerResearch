import openpyxl

wb = openpyxl.Workbook()
ws = wb.active

row = 2
intercept_row = 2
timeout_row = 2
other_row = 2
noscan_row = 2


def initialize():
    ws['A1'] = "Results"
    ws['B1'] = "HTML / Link Before"
    ws['C1'] = "HTML / Link After"
    ws['D1'] = "Tries"
    ws['G1'] = "Links with Intercept Errors:"
    ws['H1'] = "Links with Timeout Errors:"
    ws['I1'] = "Links that could not be scanned"
    ws['J1'] = "Links with unknown Errors:"


def write_intercepts(lst):
    global intercept_row
    ws[f'G{intercept_row}'] = lst
    intercept_row += 1


def write_timeout_row(lst):
    global timeout_row
    ws[f'H{timeout_row}'] = lst
    timeout_row += 1


def write_noscan_row(lst):
    global noscan_row
    ws[f'I{noscan_row}'] = lst
    noscan_row += 1

def write_other_row(lst):
    global other_row
    ws[f'J{other_row}'] = lst
    other_row += 1


def write_results(data):
    global row

    # Format = ['True redirect', HTML / link Before, HTML / link After, tries]
    if type(data) is str:
        row += 1
        ws[f'A{row}'] = data                          #website
        row += 1

    elif len(data) == 2:
        ws[f'A{row}'] = data[0]
        ws[f'B{row}'] = data[1]
        row += 1

    else:
        ws[f'A{row}'] = data[0]                       #result
        ws[f'B{row}'] = data[1]                       #before
        ws[f'C{row}'] = data[2]                       #after
        ws[f'D{row}'] = f'Tries: {data[3]}'           #number of Tries
        row += 1



    # write_intercepts(intercept)
    # write_timeout_row(timeout)
    # write_other_row(other)


def end():
    wb.save("TESTING.xlsx")

initialize()
