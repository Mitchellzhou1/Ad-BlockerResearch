import openpyxl


def write_data_to_file(data, intercept, timeout, other):
    def write_rows(lst, column, row):
        for i in range(len(lst)):
            ws[f'{column}{i+row}'] = lst[i]

    def set_up_errors(intercept, timeout, other):
        ws['A1'] = "Intercept Error"
        ws['B1'] = "Timeout Error"
        ws['c1'] = "Some Other Error"
        write_rows(intercept, 'A', 2)
        write_rows(timeout, 'B',2)
        write_rows(other, 'C', 2)
        return max(len(intercept), len(timeout), len(other))

    wb = openpyxl.Workbook()
    ws = wb.active
    row = set_up_errors(intercept, timeout, other) + 4
    ws[f'A{row}'] = "Results"
    ws[f'B{row}'] = "HTML / Link Before"
    ws[f'C{row}'] = "HTML / Link After"

    offset = row + 1
    # ['True redirect', HTML / link Before, HTML / link After, tries]

    for i in range(len(data)):
        if len(data[i]) == 1:
            ws[f'A{offset}'] = data[i][0]                       #website
            offset += 1
        else:
            ws[f'A{offset}'] = data[i][0]                       #result
            ws[f'B{offset}'] = data[i][1]                       #before
            ws[f'C{offset}'] = data[i][2]                       #after
            ws[f'D{offset}'] = f'Tries: {data[i][3]}'           #number of Tries
            offset += 1


    wb.save("TESTING.xlsx")


data_to_write = [
    ["site1"],
    ['True redirect', "HTML / link Before", "HTML / link After", "tries"],
    ['False redirect', "HTML / link Before", "HTML / link After", "tries"],
    ['True redirect', "HTML / link Before", "HTML / link After", "tries"],
    ["site2"],
    ['True redirect', "HTML / link Before", "HTML / link After", "tries"],
    ['True redirect', "HTML / link Before", "HTML / link After", "tries"]
]
intercept = ["Applice.com"]
timeout = ["Applice.com"]
other = ["dfdsa.com"]

write_data_to_file(data_to_write, intercept, timeout, other)
